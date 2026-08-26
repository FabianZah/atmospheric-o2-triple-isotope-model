"""Publication-grade XLSX export for constrained coordinate inference."""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from young_global_o2_budget import GLOBAL_MAJOR_O2_MOLES_1PAL


TITLE_FILL = PatternFill("solid", fgColor="123238")
HEADER_FILL = PatternFill("solid", fgColor="006F71")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _unit(coordinate: str) -> str:
    return {"pCO2": "ppm", "GPP": "PgC yr-1", "pO2": "PAL"}[coordinate]


def _prepare_sheet(
    sheet: Worksheet,
    *,
    title: str,
    headers: tuple[str, ...],
) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = sheet.cell(1, 1, title)
    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(3, column, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A4"


def _finish_table(sheet: Worksheet, widths: tuple[float, ...]) -> None:
    sheet.auto_filter.ref = f"A3:{sheet.cell(3, len(widths)).column_letter}{sheet.max_row}"
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _append_summary_rows(
    sheet: Worksheet,
    envelope: dict[str, Any],
    context: dict[str, Any],
) -> None:
    result = envelope["result"]
    inputs = result["inputs"]
    low, high = result["equal_tailed_credible_interval"]
    coordinate = result["solve_for"]
    unit = _unit(coordinate)
    rows: list[tuple[str, Any, str]] = [
        ("generated_utc", datetime.now(timezone.utc).isoformat(), ""),
        ("publication_model_id", envelope["publication_model_id"], ""),
        ("api_version", envelope["api_version"], ""),
        ("calculation", envelope["calculation"], ""),
        ("isotope_source", context["isotope_source"], ""),
        (
            "target_air_Delta_prime_17O_0.528",
            inputs["target_air_cap_delta17_permil"],
            "permil",
        ),
        (
            "Delta_prime_17O_analytical_sigma",
            inputs["measurement_sigma_permil"],
            "permil (1 sigma)",
        ),
        (
            "target_air_delta18O_VSMOW",
            inputs.get("target_air_delta18_conventional_permil"),
            "permil",
        ),
        (
            "delta18O_analytical_sigma",
            inputs.get("delta18_measurement_sigma_permil"),
            "permil (1 sigma)",
        ),
        ("solved_coordinate", coordinate, ""),
        ("posterior_median", result["posterior_median"], unit),
        ("credible_interval_lower", low, unit),
        ("credible_interval_upper", high, unit),
        ("credible_interval_mass", inputs["credible_mass"], "probability"),
        ("boundary_sensitive", result["solve_boundary_sensitive"], ""),
        ("surface_data_id", result["surface_data_id"], ""),
        ("upstream_model_data_id", result["upstream_model_data_id"], ""),
        (
            "central_model_data_id",
            envelope["provenance"]["central_model_data_id"],
            "",
        ),
        (
            "uncertainty_contract_id",
            envelope["provenance"]["uncertainty_contract_id"],
            "",
        ),
        ("probability_scope", result["probability_scope"], ""),
    ]
    spherule = context.get("spherule")
    if spherule:
        rows.extend(
            [
                (
                    "spherule_Delta_prime_17O_0.528",
                    spherule["cap_delta17_permil"],
                    "permil",
                ),
                (
                    "spherule_Delta_prime_17O_analytical_sigma",
                    spherule["cap_delta17_sigma_permil"],
                    "permil (1 sigma)",
                ),
                ("spherule_delta18O_VSMOW", spherule["delta18_permil"], "permil"),
                (
                    "spherule_delta18O_analytical_sigma",
                    spherule["delta18_sigma_permil"],
                    "permil (1 sigma)",
                ),
            ]
        )
    for name, constraint in inputs["constraints"].items():
        rows.append((f"{name}_constraint_kind", constraint["kind"], ""))
        for key in ("center", "sigma", "lower", "upper"):
            value = constraint.get(key)
            if value is not None:
                rows.append((f"{name}_constraint_{key}", value, _unit(name)))
        effective = result["effective_constraint_bounds"].get(name)
        if effective:
            rows.extend(
                [
                    (f"{name}_effective_lower", effective[0], _unit(name)),
                    (f"{name}_effective_upper", effective[1], _unit(name)),
                ]
            )
    for row in rows:
        sheet.append(row)


def build_coordinate_inference_workbook(
    envelope: dict[str, Any],
    context: dict[str, Any],
) -> bytes:
    """Return a self-contained XLSX export for one constrained inference."""

    result = envelope["result"]
    workbook = Workbook()
    workbook.properties.creator = "OXYTIB"
    workbook.properties.title = "Constrained atmospheric O2 isotope inference"
    workbook.properties.subject = envelope["publication_model_id"]

    summary = workbook.active
    summary.title = "Summary"
    _prepare_sheet(summary, title="Constrained model solution", headers=("Field", "Value", "Unit or note"))
    _append_summary_rows(summary, envelope, context)
    _finish_table(summary, (42, 70, 24))

    posterior = workbook.create_sheet("Posterior")
    coordinate = result["solve_for"]
    _prepare_sheet(
        posterior,
        title=f"{coordinate} marginal posterior",
        headers=(coordinate, "Unit", "Probability mass", "Probability density"),
    )
    for axis, mass, density in zip(
        result["solve_axis"],
        result["solve_marginal_probability_mass"],
        result["solve_marginal_density"],
        strict=True,
    ):
        posterior.append((axis, _unit(coordinate), mass, density))
    for row in posterior.iter_rows(min_row=4, min_col=1, max_col=4):
        row[0].number_format = "0.0000000000"
        row[2].number_format = "0.0000000000E+00"
        row[3].number_format = "0.0000000000E+00"
    _finish_table(posterior, (22, 16, 22, 22))

    if result.get("field_probability_mass") is not None:
        field = workbook.create_sheet("Joint probability")
        x_name = result["field_x_coordinate"]
        y_name = result["field_y_coordinate"]
        _prepare_sheet(
            field,
            title=f"{x_name}-{y_name} joint probability field",
            headers=(
                x_name,
                f"{x_name} unit",
                y_name,
                f"{y_name} unit",
                "Probability mass",
                "Probability density",
                "Inside 95% HPD",
            ),
        )
        y_axis = result["field_y_axis"]
        for x_index, x_value in enumerate(result["field_x_axis"]):
            for y_index, y_value in enumerate(y_axis):
                flat_index = x_index * len(y_axis) + y_index
                field.append(
                    (
                        x_value,
                        _unit(x_name),
                        y_value,
                        _unit(y_name),
                        result["field_probability_mass"][flat_index],
                        result["field_density"][flat_index],
                        result["field_hpd_mask"][flat_index],
                    )
                )
        for row in field.iter_rows(min_row=4, min_col=1, max_col=7):
            row[0].number_format = "0.0000000000"
            row[2].number_format = "0.0000000000"
            row[4].number_format = "0.0000000000E+00"
            row[5].number_format = "0.0000000000E+00"
        _finish_table(field, (20, 16, 20, 16, 22, 22, 18))

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=True, sort_keys=True)


def _flatten_mapping(value: Any, prefix: str = "") -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = []
    if isinstance(value, dict):
        for key, child in value.items():
            path = f"{prefix}.{key}" if prefix else str(key)
            rows.extend(_flatten_mapping(child, path))
        return rows
    rows.append((prefix, _cell_value(value)))
    return rows


def build_transient_workbook(
    envelope: dict[str, Any],
    experiment_type: str,
) -> bytes:
    """Return an XLSX record of a public time-response experiment."""

    result = envelope["result"]
    request = result["request"]
    equilibrium = result.get("operational_equilibrium", {})
    workbook = Workbook()
    workbook.properties.creator = "OXYTIB"
    workbook.properties.title = "Atmospheric O2 isotope time-response experiment"
    workbook.properties.subject = envelope["publication_model_id"]

    summary = workbook.active
    summary.title = "Summary"
    _prepare_sheet(
        summary,
        title="Atmospheric isotope time response",
        headers=("Field", "Value", "Unit or note"),
    )
    summary_rows = [
        ("generated_utc", datetime.now(timezone.utc).isoformat(), ""),
        ("publication_model_id", envelope["publication_model_id"], ""),
        ("api_version", envelope["api_version"], ""),
        ("calculation", envelope["calculation"], ""),
        ("experiment_type", experiment_type, ""),
        ("display_duration", request["duration_years"], "years"),
        ("sample_count", request["sample_count"], ""),
        (
            "equilibrium_search_horizon",
            request["equilibrium_search_max_years"],
            "years",
        ),
        (
            "operational_equilibrium_time",
            equilibrium.get("time_years"),
            "years",
        ),
        (
            "operational_equilibrium_tolerance",
            equilibrium.get("tolerance_permil"),
            "permil",
        ),
        ("model_data_id", result.get("model_data_id"), ""),
        ("transfer_convention", result.get("transfer_convention"), ""),
        ("carbon_driver_preset", result.get("carbon_driver_preset"), ""),
    ]
    for row in summary_rows:
        summary.append(row)
    _finish_table(summary, (42, 72, 24))

    inputs = workbook.create_sheet("Inputs")
    _prepare_sheet(
        inputs,
        title="Experiment inputs",
        headers=("Parameter", "Value"),
    )
    for path, value in _flatten_mapping(request):
        inputs.append((path, value))
    _finish_table(inputs, (48, 72))

    timeseries = workbook.create_sheet("Time series")
    headers = (
        "time_years",
        "pO2_PAL",
        "pCO2_ppm",
        "GPP_PgC_per_year",
        "photosynthesis_fraction_of_initial",
        "carbon_driver_pO2_PAL",
        "O2_Delta_prime_17O_0.528_permil",
        "O2_delta_prime_18O_permil",
        "O16O16_mol",
        "O16O17_mol",
        "O16O18_mol",
    )
    _prepare_sheet(timeseries, title="Model time series", headers=headers)
    states = result["states"]
    is_photosynthesis = experiment_type == "photosynthesis"
    is_trajectory = experiment_type == "pCO2_trajectory"
    if is_photosynthesis:
        pco2_values = result["pco2_ppm"]
        carbon_po2_values = result["carbon_driver_po2_pal"]
        gpp_value = request["initial"]["gpp_pgC_per_year"]
        photosynthesis_fraction = request["photosynthesis_fraction"]
    elif is_trajectory:
        pco2_values = result["pco2_ppm"]
        carbon_po2_values = [None] * len(states)
        gpp_value = request["initial"]["gpp_pgC_per_year"]
        photosynthesis_fraction = None
    else:
        pco2_values = [request["final"]["p_co2_ppm"]] * len(states)
        carbon_po2_values = [None] * len(states)
        gpp_value = request["final"]["gpp_pgC_per_year"]
        photosynthesis_fraction = None
    for index, (time, state) in enumerate(
        zip(result["time_years"], states, strict=True)
    ):
        timeseries.append(
            (
                time,
                state["o16o16_mol"] / GLOBAL_MAJOR_O2_MOLES_1PAL,
                pco2_values[index],
                gpp_value,
                photosynthesis_fraction,
                carbon_po2_values[index],
                state["cap_delta17_prime_permil"],
                state["delta18_prime_permil"],
                state["o16o16_mol"],
                state["o16o17_mol"],
                state["o16o18_mol"],
            )
        )
    for row in timeseries.iter_rows(min_row=4, max_col=len(headers)):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.0000000000E+00"
    _finish_table(
        timeseries,
        (16, 16, 18, 22, 28, 22, 30, 27, 22, 22, 22),
    )

    provenance = workbook.create_sheet("Provenance")
    _prepare_sheet(
        provenance,
        title="Model and solver provenance",
        headers=("Field", "Value"),
    )
    provenance_values = {
        **envelope.get("provenance", {}),
        "solver": result.get("solver", {}),
        "operational_equilibrium": equilibrium,
    }
    for path, value in _flatten_mapping(provenance_values):
        provenance.append((path, value))
    _finish_table(provenance, (52, 88))

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
