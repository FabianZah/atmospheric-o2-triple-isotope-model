"""Public HTTP API contract tests."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

from spherule_to_air_d17o import air_d17o_from_spherule
from public_model_service import to_jsonable
import web_api
from web_api import app
from web_api import _cors_origins


client = TestClient(app)


def test_frontend_assets_and_api_work_when_mounted_below_prefix() -> None:
    prefixed = FastAPI()
    prefixed.mount("/oxytib", app)
    prefixed_client = TestClient(prefixed)

    root = prefixed_client.get("/oxytib/")
    script = prefixed_client.get("/oxytib/assets/app.js")
    styles = prefixed_client.get("/oxytib/assets/styles.css")
    health = prefixed_client.get("/oxytib/api/v1/health")
    citation = prefixed_client.get("/oxytib/citation/model.bib")
    documentation = prefixed_client.get("/oxytib/docs")
    swagger_initializer = prefixed_client.get("/oxytib/assets/swagger-init.js")

    assert root.status_code == 200
    assert 'href="assets/styles.css?v=1.20.6"' in root.text
    assert 'src="assets/app.js?v=1.20.6"' in root.text
    assert script.status_code == 200
    assert styles.status_code == 200
    assert health.status_code == 200
    assert health.json()["status"] == "ok"
    assert citation.status_code == 200
    assert documentation.status_code == 200
    assert 'src="assets/swagger-init.js"' in documentation.text
    assert "SwaggerUIBundle" not in documentation.text
    assert swagger_initializer.status_code == 200
    assert 'replace(/\\/docs\\/?$/, "")' in swagger_initializer.text
    assert 'url: `${applicationPrefix}/openapi.json`' in swagger_initializer.text


def test_cors_is_opt_in(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.delenv("OXYTIB_CORS_ORIGINS", raising=False)
    assert _cors_origins() == []
    monkeypatch.setenv(
        "OXYTIB_CORS_ORIGINS",
        "https://model.example.org, https://analysis.example.org",
    )
    assert _cors_origins() == [
        "https://model.example.org",
        "https://analysis.example.org",
    ]


def test_public_responses_include_browser_security_headers() -> None:
    response = client.get("/")
    assert response.status_code == 200
    assert response.headers["x-content-type-options"] == "nosniff"
    assert response.headers["x-frame-options"] == "DENY"
    assert response.headers["referrer-policy"] == "no-referrer"
    assert response.headers["permissions-policy"] == (
        "camera=(), microphone=(), geolocation=()"
    )
    assert "frame-ancestors 'none'" in response.headers["content-security-policy"]


def test_oversized_request_body_is_rejected_before_model_execution() -> None:
    response = client.post(
        "/api/v1/forward",
        content=b"x" * (web_api.MAX_REQUEST_BYTES + 1),
        headers={"Content-Type": "application/json"},
    )
    assert response.status_code == 413
    assert response.json()["detail"] == "request body exceeds the public API size limit"


def test_root_serves_independent_frontend_and_static_assets() -> None:
    root = client.get("/")
    assert root.status_code == 200
    assert '<h1 aria-label="OXYTIB">' in root.text
    assert "streamlit" not in root.text.lower()
    assert 'type="number"' not in root.text
    assert 'id="transient-forcing"' in root.text
    assert 'id="transient-pco2"' in root.text
    assert root.text.count('class="plot-export"') == 7
    assert 'id="download-result-xlsx"' in root.text
    assert 'id="download-result"' not in root.text
    assert 'id="download-transient-xlsx"' in root.text
    assert 'id="transient-progress"' in root.text
    assert 'id="transient-progress-elapsed"' in root.text
    assert 'id="solver-progress"' in root.text
    assert 'id="solver-progress-elapsed"' in root.text
    assert 'href="assets/styles.css?v=1.20.6"' in root.text
    assert 'src="assets/app.js?v=1.20.6"' in root.text
    assert '>Download XLSX</button>' in root.text
    assert 'class="brand-lockup"' in root.text
    assert '<span>OXY</span><strong>TIB</strong>' in root.text
    assert (
        "O<sub>2</sub> · Δ′<sup>17</sup>O · pCO<sub>2</sub> · "
        "pO<sub>2</sub> · GPP"
    ) in root.text
    assert 'id="theme-toggle"' in root.text
    assert 'href="/assets/' not in root.text
    assert 'src="/assets/' not in root.text
    assert "x ∈ {17, 18}" in root.text
    assert root.text.index("δ<sup>x</sup>O = 1000") < root.text.index("δ′<sup>x</sup>O = 1000")
    assert "Present atmospheric level" in root.text
    assert "1.00 PAL corresponds to 21.2% atmospheric O<sub>2</sub>" in root.text
    assert "4.18 × 10<sup>19</sup> mol" not in root.text
    assert "Why is the pCO<sub>2</sub> axis logarithmic?" not in root.text
    assert "Sources are grouped by how they inform the model" not in root.text
    assert "Why must two variables be constrained before solving for the third?" in root.text
    assert "What is the difference between the isotope field and the constrained solution?" in root.text
    assert "What assumptions are made about input uncertainties?" in root.text
    assert "What does a boundary-limited solution mean?" in root.text
    assert "Does the main solver represent steady state?" in root.text
    assert "How should the time-response experiments be interpreted?" in root.text
    assert "How long do calculations take?" in root.text
    assert "Can the model be applied to ancient atmospheres?" in root.text
    assert "How should cosmic-spherule uncertainties be entered?" in root.text
    assert "remain within 0.001‰ of the long-run state" in root.text
    assert "All Δ" not in root.text
    assert 'id="air-d18"' in root.text
    assert 'for="air-sigma">Δ′<sup>17</sup>O 1σ' in root.text
    assert 'for="spherule-d17-sigma">Δ′<sup>17</sup>O 1σ' in root.text
    assert 'id="spherule-d17-sigma" type="text" inputmode="decimal" value="0.060"' in root.text
    assert 'id="spherule-d18-sigma" type="text" inputmode="decimal" value="0.500"' in root.text
    assert 'id="surface-po2"' in root.text
    assert 'id="surface-po2-readout"' not in root.text
    assert 'id="pco2-constraint-mode"' in root.text
    assert 'id="gpp-constraint-mode"' in root.text
    assert 'id="po2-constraint-mode"' in root.text
    assert "Δ′<sup>17</sup>O<sub>0.528</sub>" in root.text
    assert "Vienna Standard Mean Ocean Water reference scale." in root.text
    assert 'class="marginal-key"' not in root.text
    assert 'id="isotope-summary"' not in root.text
    assert "https://doi.org/10.1016/j.gca.2014.03.026" in root.text
    assert "How to cite" in root.text
    assert 'href="citation/model.bib"' in root.text
    assert 'href="citation/model.ris"' in root.text
    assert 'href="citation/CITATION.cff"' in root.text
    assert '<a href="docs">API documentation</a>' in root.text
    assert 'href="/docs"' not in root.text
    assert root.text.count('class="reference-group"') == 5
    assert root.text.count("<article>", root.text.index('id="view-references"')) == 24
    for citation_url in (
        "https://jpldataeval.jpl.nasa.gov/",
        "https://doi.org/10.1002/qj.3803",
        "https://doi.org/10.3847/PSJ/ae0e1c",
        "https://doi.org/10.1029/2003GL018451",
        "https://doi.org/10.5194/amt-18-2701-2025",
        "https://doi.org/10.1126/science.abj8826",
        "https://doi.org/10.1098/rspb.1999.0852",
            "https://doi.org/10.1146/annurev-earth-032320-095425",
            "https://www.ipcc.ch/report/ar6/wg1/chapter/chapter-2/",
            "https://gml.noaa.gov/aggi/aggi.html",
    ):
        assert citation_url in root.text
    assert "<i>Scientific Reports</i>, 13, 2162." in root.text

    script = client.get("/assets/app.js")
    styles = client.get("/assets/styles.css")
    assert script.status_code == 200
    assert styles.status_code == 200
    assert "/api/v1/inference/coordinate" in script.text
    assert "/api/v1/export/coordinate.xlsx" in script.text
    assert "/api/v1/export/transient.xlsx" in script.text
    assert "/api/v1/field/isotope" in script.text
    assert "const APPLICATION_BASE_PATH" in script.text
    assert "function applicationUrl" in script.text
    assert "function beginCalculationProgress" in script.text
    assert "function beginTransientProgress" in script.text
    assert "function beginSolverProgress" in script.text
    assert "performance.now() - startedAt" in script.text
    assert "fetch(applicationUrl(path)" in script.text
    assert 'fetch(applicationUrl("/api/v1/export/coordinate.xlsx")' in script.text
    assert "Deterministic model isotope field" not in script.text
    assert "contour is emphasized" not in script.text
    assert "level === -10" not in script.text
    assert "result.contour_levels_permil" in script.text
    assert "result.contour_label_decimals" in script.text
    assert "function adaptiveContourLevels" not in script.text
    assert "function fieldBalancedContourLevels" not in script.text
    assert "canvas.dataset.contourValues" in script.text
    assert "balanced across the plotted field" in script.text
    assert "level === -24 || level === -26" not in script.text
    assert "Color encodes the model-predicted atmospheric" not in script.text
    assert "drawMarginalLegend(ctx, width, edgeLimited)" in script.text
    assert '"95% credible region"' in script.text
    assert '"Atmospheric O₂ Δ′¹⁷O₀.₅₂₈ (‰)"' in script.text
    assert "Fixed pO₂ =" in script.text
    assert "Peak grid compatibility" not in script.text
    assert '"No interior solution"' in script.text
    assert "do not identify an interior" in script.text
    assert "solve_boundary_probability_mass >= 0.5" in script.text
    assert '"Relative compatibility"' in script.text
    assert '"Domain-truncated 95% interval"' in script.text
    assert "function drawMarginalPosterior" in script.text
    assert 'prior: state.solveFor === "pO2" ? "uniform" : "log_uniform"' not in script.text
    assert "pco2_grid_size: 241" in script.text
    assert "gpp_grid_size: 201" in script.text
    assert "function validateSurfaceDomain" in script.text
    assert "GPP limits must remain within" in script.text
    assert "function publicErrorMessage" in script.text
    assert "GPP must remain within" in script.text
    assert 'state.source = "air"' in script.text
    assert 'state.solveFor = "pCO2"' in script.text
    assert "result.pco2_ppm" in script.text
    assert 'formatFixed(d17[0], 3)' in script.text
    assert "At transition end Δ′" in script.text
    assert "At transition end δ′" in script.text
    assert "displayTimes = [-preStepDuration" in script.text
    assert "function niceTickStep" in script.text
    assert "Math.ceil(xmin / xTickStep)" in script.text
    assert 'canvas.toBlob' in script.text
    assert "--teal" in styles.text


@pytest.mark.parametrize(
    ("path", "marker"),
    [
        ("/citation/model.bib", "@software{zahnow_oxytib_2026"),
        ("/citation/model.ris", "TY  - COMP"),
        ("/citation/CITATION.cff", "cff-version: 1.2.0"),
    ],
)
def test_downloadable_citation_records(path: str, marker: str) -> None:
    response = client.get(path)
    assert response.status_code == 200
    assert marker in response.text
    assert "attachment" in response.headers["content-disposition"]


def test_coordinate_xlsx_export_contains_provenance_and_posterior_data() -> None:
    response = client.post(
        "/api/v1/export/coordinate.xlsx",
        json={
            "inference": {
                "solve_for": "pCO2",
                "target_air_cap_delta17_permil": -0.432,
                "measurement_sigma_permil": 0.015,
                "target_air_delta18_conventional_permil": 23.9,
                "delta18_measurement_sigma_permil": 0.3,
                "gpp_constraint": {"kind": "fixed", "center": 290.0},
                "po2_constraint": {"kind": "fixed", "center": 1.0},
                "pco2_grid_size": 17,
                "gpp_grid_size": 17,
                "po2_grid_size": 17,
            },
            "context": {"isotope_source": "Direct air O2"},
        },
    )

    assert response.status_code == 200
    assert response.content[:2] == b"PK"
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "oxytib_pco2_solution.xlsx" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook.sheetnames == ["Summary", "Posterior"]
    summary = {
        row[0].value: row[1].value
        for row in workbook["Summary"].iter_rows(min_row=4, max_col=2)
    }
    assert summary["software"] == "OXYTIB"
    assert summary["software_version"] == "0.1.0"
    assert "publication_model_id" not in summary
    assert "surface_data_id" not in summary
    assert summary["isotope_source"] == "Direct air O2"
    assert summary["solved_coordinate"] == "pCO2"
    assert summary["GPP_constraint_kind"] == "fixed"
    assert workbook["Posterior"].max_row == 20


def test_transient_xlsx_export_reuses_run_and_contains_metadata(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    web_api._cached_state_step.cache_clear()
    original = web_api.state_step_transient
    calls = 0

    def counted(request):
        nonlocal calls
        calls += 1
        return original(request)

    monkeypatch.setattr(web_api, "state_step_transient", counted)
    state = {"p_o2_pal": 1.0, "p_co2_ppm": 294.0, "gpp_pgC_per_year": 290.0}
    experiment = {
        "initial": state,
        "final": state,
        "duration_years": 10.0,
        "sample_count": 3,
        "equilibrium_search_max_years": 10.0,
    }
    run = client.post("/api/v1/transients/state-step", json=experiment)
    export = client.post(
        "/api/v1/export/transient.xlsx",
        json={"experiment_type": "pCO2", "state_step": experiment},
    )

    assert run.status_code == 200
    assert export.status_code == 200
    assert calls == 1
    assert export.content[:2] == b"PK"
    assert "oxytib_pco2_time_response.xlsx" in export.headers[
        "content-disposition"
    ]
    workbook = load_workbook(BytesIO(export.content), data_only=False)
    assert workbook.sheetnames == ["Summary", "Inputs", "Time series", "Provenance"]
    summary = {
        row[0].value: row[1].value
        for row in workbook["Summary"].iter_rows(min_row=4, max_col=2)
    }
    assert summary["software"] == "OXYTIB"
    assert summary["software_version"] == "0.1.0"
    assert "publication_model_id" not in summary
    provenance_fields = {
        row[0].value
        for row in workbook["Provenance"].iter_rows(min_row=4, max_col=1)
    }
    assert not any(field and field.endswith("_id") for field in provenance_fields)
    assert summary["experiment_type"] == "pCO2"
    assert summary["sample_count"] == 3
    assert "transfer_convention" not in summary
    assert "carbon_driver_preset" not in summary
    assert workbook["Time series"].max_row == 6


def test_gradual_pco2_trajectory_endpoint_and_xlsx_export() -> None:
    experiment = {
        "initial": {
            "p_o2_pal": 1.0,
            "p_co2_ppm": 285.5,
            "gpp_pgC_per_year": 290.0,
        },
        "final_pco2_ppm": 422.8,
        "transition_duration_years": 174.0,
        "interpolation": "smoothstep",
        "duration_years": 200.0,
        "sample_count": 9,
        "equilibrium_search_max_years": 200.0,
    }
    run = client.post("/api/v1/transients/pco2-trajectory", json=experiment)
    export = client.post(
        "/api/v1/export/transient.xlsx",
        json={
            "experiment_type": "pCO2_trajectory",
            "pco2_trajectory": experiment,
        },
    )

    assert run.status_code == 200
    body = run.json()
    assert body["calculation"] == "pco2_trajectory_transient"
    assert body["result"]["pco2_ppm"][0] == pytest.approx(285.5)
    assert body["result"]["pco2_ppm"][-1] == pytest.approx(422.8)
    assert body["result"]["transition_end_state"]["time_years"] == pytest.approx(
        174.0
    )
    assert export.status_code == 200
    assert export.content[:2] == b"PK"
    assert "oxytib_pco2_trajectory_time_response.xlsx" in export.headers[
        "content-disposition"
    ]
    workbook = load_workbook(BytesIO(export.content), data_only=False)
    summary = {
        row[0].value: row[1].value
        for row in workbook["Summary"].iter_rows(min_row=4, max_col=2)
    }
    assert summary["experiment_type"] == "pCO2_trajectory"
    assert summary["transition_end_time"] == pytest.approx(174.0)
    assert "transition_end_O2_Delta_prime_17O_0.528" in summary
    assert workbook["Time series"].max_row == 12


def test_coordinate_xlsx_export_includes_spherule_context_and_joint_field() -> None:
    response = client.post(
        "/api/v1/export/coordinate.xlsx",
        json={
            "inference": {
                "solve_for": "pCO2",
                "target_air_cap_delta17_permil": -0.432,
                "measurement_sigma_permil": 0.062,
                "gpp_constraint": {"kind": "normal", "center": 290.0, "sigma": 29.0},
                "po2_constraint": {"kind": "fixed", "center": 1.0},
                "pco2_grid_size": 17,
                "gpp_grid_size": 17,
                "po2_grid_size": 17,
            },
            "context": {
                "isotope_source": "I-type cosmic spherule",
                "spherule": {
                    "cap_delta17_permil": -0.66,
                    "cap_delta17_sigma_permil": 0.06,
                    "delta18_permil": 43.269,
                    "delta18_sigma_permil": 0.5,
                },
            },
        },
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook.sheetnames == ["Summary", "Posterior", "Joint probability"]
    summary = {
        row[0].value: row[1].value
        for row in workbook["Summary"].iter_rows(min_row=4, max_col=2)
    }
    assert summary["isotope_source"] == "I-type cosmic spherule"
    assert summary["spherule_delta18O_VSMOW"] == pytest.approx(43.269)
    assert workbook["Joint probability"].max_row == 292


def test_health_and_model_metadata_expose_one_accepted_model() -> None:
    health = client.get("/api/v1/health")
    assert health.status_code == 200
    assert health.json() == {
        "status": "ok",
        "api_version": "1.0",
        "publication_model_id": "oxytib_publication_model_v1",
    }

    metadata = client.get("/api/v1/model")
    assert metadata.status_code == 200
    body = metadata.json()
    assert body["operational_domain"]["pCO2_ppm"]["minimum"] == 50.0
    assert body["operational_domain"]["pCO2_ppm"]["maximum"] == 60000.0
    assert body["uncertainty"]["layers_remain_separate"] is True
    assert body["citation"]["citation_files"] == [
        "CITATION.cff",
        "CITATION.bib",
        "CITATION.ris",
    ]
    assert body["deterministic_model"]["numerical_accelerator"][
        "extrapolation_permitted"
    ] is False


def test_forward_and_inverse_roundtrip_preserve_model_provenance() -> None:
    forward = client.post(
        "/api/v1/forward",
        json={"p_o2_pal": 1.0, "p_co2_ppm": 294.0, "gpp_pgC_per_year": 290.0},
    )
    assert forward.status_code == 200
    forward_body = forward.json()
    target = forward_body["result"]["central_cap_delta17_prime_permil"]
    assert target == pytest.approx(-0.42635313046373885)
    assert forward_body["provenance"]["extrapolation_permitted"] is False

    inverse = client.post(
        "/api/v1/inverse",
        json={
            "target_air_cap_delta17_permil": target,
            "solve_for": "pCO2",
            "measurement_uncertainty_permil": 0.0,
            "p_o2_pal": 1.0,
            "gpp_pgC_per_year": 290.0,
        },
    )
    assert inverse.status_code == 200
    inverse_body = inverse.json()
    assert inverse_body["result"]["central_root"] == pytest.approx(294.0)
    assert inverse_body["result"]["live_root_verified"] is True
    assert inverse_body["publication_model_id"] == forward_body["publication_model_id"]


def test_spherule_conversion_propagates_analytical_uncertainty() -> None:
    response = client.post(
        "/api/v1/proxy/spherule-to-air",
        json={
            "cap_delta17_spherule_permil": -10.0,
            "delta18_spherule_permil": 42.0,
            "cap_delta17_sigma_permil": 0.07,
            "delta18_sigma_permil": 0.2,
        },
    )
    assert response.status_code == 200
    result = response.json()["result"]
    assert result["cap_delta17_air_o2_permil"] == pytest.approx(
        air_d17o_from_spherule(-10.0, 42.0)
    )
    assert result["analytical_sigma_permil"] > 0.07
    assert result["calibration_sensitivity_is_confidence_interval"] is False


def test_joint_posterior_returns_summary_by_default_and_grid_on_request() -> None:
    request = {
        "target_air_cap_delta17_permil": -1.0,
        "measurement_sigma_permil": 0.02,
        "free_coordinates": ["pCO2", "GPP"],
        "p_o2_pal": 1.0,
        "pco2_grid_size": 17,
        "gpp_grid_size": 17,
    }
    summary = client.post("/api/v1/posterior/joint", json=request)
    assert summary.status_code == 200
    summary_result = summary.json()["result"]
    assert summary_result["posterior_shape"] == [17, 17]
    assert summary_result["grid_included"] is False
    assert "posterior_probability_mass" not in summary_result
    assert set(summary_result["equal_tailed_credible_intervals"]) == {"pCO2", "GPP"}

    grid = client.post(
        "/api/v1/posterior/joint?include_grid=true", json=request
    )
    assert grid.status_code == 200
    grid_result = grid.json()["result"]
    assert len(grid_result["posterior_probability_mass"]) == 17 * 17
    assert len(grid_result["hpd_mask"]) == 17 * 17


def test_constrained_pco2_endpoint_propagates_gpp_and_po2_constraints() -> None:
    response = client.post(
        "/api/v1/inference/pco2",
        json={
            "target_air_cap_delta17_permil": -0.432,
            "measurement_sigma_permil": 0.015,
            "target_air_delta18_conventional_permil": 23.9,
            "delta18_measurement_sigma_permil": 0.3,
            "gpp_constraint": {"kind": "normal", "center": 290.0, "sigma": 29.0},
            "po2_constraint": {"kind": "range", "lower": 0.9, "upper": 1.1},
            "pco2_grid_size": 61,
            "gpp_grid_size": 41,
            "po2_grid_size": 17,
        },
    )
    assert response.status_code == 200
    result = response.json()["result"]
    assert result["pco2_gpp_shape"] == [61, 41]
    assert sum(result["pco2_marginal_probability_mass"]) == pytest.approx(1.0)
    assert sum(result["pco2_gpp_probability_mass"]) == pytest.approx(1.0)
    assert "No log-uniform" in result["probability_scope"]


def test_coordinate_inference_endpoint_accepts_constraints_for_any_solved_axis() -> None:
    response = client.post(
        "/api/v1/inference/coordinate",
        json={
            "solve_for": "GPP",
            "target_air_cap_delta17_permil": -0.432,
            "measurement_sigma_permil": 0.015,
            "pco2_constraint": {"kind": "range", "lower": 250.0, "upper": 400.0},
            "po2_constraint": {"kind": "normal", "center": 1.0, "sigma": 0.05},
            "pco2_grid_size": 61,
            "gpp_grid_size": 41,
            "po2_grid_size": 17,
        },
    )
    assert response.status_code == 200
    result = response.json()["result"]
    assert result["solve_for"] == "GPP"
    assert result["field_x_coordinate"] == "pCO2"
    assert result["field_y_coordinate"] == "GPP"
    assert sum(result["solve_marginal_probability_mass"]) == pytest.approx(1.0)


@pytest.mark.parametrize("solve_for", ["pCO2", "GPP", "pO2"])
@pytest.mark.parametrize("first_kind", ["fixed", "normal", "range"])
@pytest.mark.parametrize("second_kind", ["fixed", "normal", "range"])
def test_coordinate_inference_public_constraint_matrix(
    solve_for: str, first_kind: str, second_kind: str
) -> None:
    definitions = {
        "pCO2": {
            "fixed": {"kind": "fixed", "center": 294.0},
            "normal": {"kind": "normal", "center": 294.0, "sigma": 20.0},
            "range": {"kind": "range", "lower": 250.0, "upper": 400.0},
        },
        "GPP": {
            "fixed": {"kind": "fixed", "center": 290.0},
            "normal": {"kind": "normal", "center": 290.0, "sigma": 20.0},
            "range": {"kind": "range", "lower": 250.0, "upper": 330.0},
        },
        "pO2": {
            "fixed": {"kind": "fixed", "center": 1.0},
            "normal": {"kind": "normal", "center": 1.0, "sigma": 0.05},
            "range": {"kind": "range", "lower": 0.9, "upper": 1.1},
        },
    }
    field_names = {
        "pCO2": "pco2_constraint",
        "GPP": "gpp_constraint",
        "pO2": "po2_constraint",
    }
    constrained = [name for name in ("pCO2", "GPP", "pO2") if name != solve_for]
    kinds = dict(zip(constrained, (first_kind, second_kind), strict=True))
    request = {
        "solve_for": solve_for,
        "target_air_cap_delta17_permil": -0.432,
        "measurement_sigma_permil": 0.015,
        "target_air_delta18_conventional_permil": 23.9,
        "delta18_measurement_sigma_permil": 0.3,
        "pco2_grid_size": 17,
        "gpp_grid_size": 17,
        "po2_grid_size": 17,
    }
    for name in constrained:
        request[field_names[name]] = definitions[name][kinds[name]]

    response = client.post("/api/v1/inference/coordinate", json=request)
    assert response.status_code == 200, response.text
    result = response.json()["result"]
    assert result["solve_for"] == solve_for
    assert sum(result["solve_marginal_probability_mass"]) == pytest.approx(1.0)
    low, high = result["equal_tailed_credible_interval"]
    assert low <= result["posterior_median"] <= high
    assert set(result["effective_constraint_bounds"]) == set(constrained)
    assert result["final_solve_axis_size"] == result["initial_solve_axis_size"]
    assert result["final_solve_bounds"][0] >= result["initial_solve_bounds"][0]
    assert result["final_solve_bounds"][1] <= result["initial_solve_bounds"][1]
    assert 0.0 <= result["solve_boundary_probability_mass"] <= 1.0
    assert isinstance(result["solve_mode_at_boundary"], bool)
    if result["solve_boundary_sensitive"]:
        assert result["solve_boundary_direction"] in {"lower", "upper"}
    else:
        assert result["solve_boundary_direction"] is None
    has_uncertain_constraint = any(kind != "fixed" for kind in kinds.values())
    assert (result["field_probability_mass"] is not None) is has_uncertain_constraint
    if has_uncertain_constraint:
        assert sum(result["field_probability_mass"]) == pytest.approx(1.0)
        assert result["field_hpd_probability_mass"] >= 0.95


def test_direct_air_delta18_contributes_to_conditional_and_joint_likelihoods() -> None:
    common = {
        "target_air_cap_delta17_permil": -0.432,
        "measurement_sigma_permil": 0.015,
        "target_air_delta18_conventional_permil": 23.9,
        "delta18_measurement_sigma_permil": 0.3,
        "p_o2_pal": 1.0,
        "gpp_pgC_per_year": 290.0,
    }
    conditional = client.post(
        "/api/v1/posterior/conditional",
        json={**common, "solve_for": "pCO2", "grid_size": 257},
    )
    assert conditional.status_code == 200
    conditional_result = conditional.json()["result"]
    assert len(conditional_result["model_delta18_conventional_permil"]) == 257
    assert conditional_result["model_delta18_at_mode_permil"] > 0.0
    assert "conventional delta-18O" in conditional_result["probability_scope"]

    joint = client.post(
        "/api/v1/posterior/joint?include_grid=true",
        json={
            **common,
            "free_coordinates": ["pCO2", "GPP"],
            "pco2_grid_size": 17,
            "gpp_grid_size": 17,
        },
    )
    assert joint.status_code == 200
    joint_result = joint.json()["result"]
    assert len(joint_result["model_delta18_conventional_permil"]) == 17 * 17
    assert "conventional delta-18O" in joint_result["probability_scope"]


def test_deterministic_isotope_field_is_separate_from_the_posterior() -> None:
    response = client.post(
        "/api/v1/field/isotope",
        json={
            "p_o2_pal": 1.0,
            "pco2_bounds_ppm": [50.0, 60000.0],
            "gpp_bounds_pgC_per_year": [18.256264, 435.0],
            "pco2_grid_size": 17,
            "gpp_grid_size": 19,
        },
    )
    assert response.status_code == 200
    result = response.json()["result"]
    assert result["field_shape"] == [17, 19]
    assert len(result["central_cap_delta17_permil"]) == 17 * 19
    assert result["minimum_cap_delta17_permil"] < result["maximum_cap_delta17_permil"]
    assert 8 <= len(result["contour_levels_permil"]) <= 10
    assert result["contour_levels_permil"] == sorted(result["contour_levels_permil"])
    assert result["contour_selection_strategy"] == "plot_area_balanced_readable"
    assert "handled by inference endpoints" in result["field_scope"]


def test_out_of_domain_and_unknown_fields_are_rejected() -> None:
    outside = client.post(
        "/api/v1/forward",
        json={"p_o2_pal": 1.0, "p_co2_ppm": 49.0, "gpp_pgC_per_year": 290.0},
    )
    assert outside.status_code == 422
    assert "outside" in outside.json()["detail"]

    extra = client.post(
        "/api/v1/forward",
        json={
            "p_o2_pal": 1.0,
            "p_co2_ppm": 294.0,
            "gpp_pgC_per_year": 290.0,
            "development_branch": "young",
        },
    )
    assert extra.status_code == 422

    oversized = client.post(
        "/api/v1/posterior/joint",
        json={
            "target_air_cap_delta17_permil": -1.0,
            "measurement_sigma_permil": 0.02,
            "free_coordinates": ["pCO2", "GPP", "pO2"],
            "pco2_grid_size": 161,
            "gpp_grid_size": 121,
            "po2_grid_size": 81,
        },
    )
    assert oversized.status_code == 422
    assert "250,000-cell" in oversized.text

    assert to_jsonable(float("nan")) is None
    assert to_jsonable(float("inf")) is None


def test_openapi_declares_transient_and_inference_routes() -> None:
    schema = client.get("/openapi.json")
    assert schema.status_code == 200
    paths = schema.json()["paths"]
    required = {
        "/api/v1/forward",
        "/api/v1/inverse",
        "/api/v1/posterior/conditional",
        "/api/v1/posterior/joint",
        "/api/v1/inference/pco2",
        "/api/v1/inference/coordinate",
        "/api/v1/field/isotope",
        "/api/v1/proxy/spherule-to-air",
        "/api/v1/transients/state-step",
        "/api/v1/transients/photosynthesis-step",
        "/api/v1/export/transient.xlsx",
    }
    assert required.issubset(paths)


def test_state_step_route_returns_time_resolved_isotope_response() -> None:
    state = {"p_o2_pal": 1.0, "p_co2_ppm": 294.0, "gpp_pgC_per_year": 290.0}
    response = client.post(
        "/api/v1/transients/state-step",
        json={
            "initial": state,
            "final": state,
            "duration_years": 10.0,
            "sample_count": 3,
            "equilibrium_search_max_years": 10.0,
        },
    )
    assert response.status_code == 200
    body = response.json()
    assert body["calculation"] == "state_step_transient"
    assert body["result"]["time_years"] == [0.0, 5.0, 10.0]
    assert len(body["result"]["states"]) == 3
    assert all("cap_delta17_prime_permil" in state for state in body["result"]["states"])
